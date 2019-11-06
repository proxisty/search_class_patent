from openpyxl import load_workbook


def read_file(name):
    """На вход подается строка элементов класса
       На выходе массив элементов класса"""
    with open(name, 'r', encoding='UTF-8') as file:
        text = file.read()
        text = text[5:-1]
        text_mass = text.split('; ')
        return text_mass


def search_index_element(text_mass):
    """Возвращаем B колонку индексов с массива и элементы с ошибками"""
    mass_index = []
    mass_without_error = []
    wb = load_workbook('classes_11_19_rus.xlsx')
    ws = wb.active
    ws.title = "classes_rus"
    for mass_value in range(len(text_mass)):
        for row in range(1, ws.max_row):
            if text_mass[mass_value].rstrip() == str(ws.cell(row=row, column=3).value).rstrip().lower().replace('*', ''):
                mass_index.append(ws.cell(row=row, column=2).value)
                mass_without_error.append(text_mass[mass_value].rstrip())
    error_elements = list(set(text_mass) - set(mass_without_error))
    return mass_index, error_elements


def search_translate(index_mass):
    """ ОШИИБКАА,ИСПААВИТЬ
    СБИВАЕТСЯ ПЕРЕВОД ИЗ-за строчек КЛАСС ...
    """
    error_index = []
    mass_translate = []
    n_error_index = []
    wb = load_workbook('classes_11_19_3lang.xlsx')
    ws = wb.active
    ws.title = "classes_3leng.xlsx"
    for index in range(len(index_mass)):
        for row in range(3, ws.max_row, 3):
            if index_mass[index] == str(ws.cell(row=row, column=2).value):
                mass_translate.append(str(ws.cell(row=row, column=3).value))
                n_error_index.append(index_mass[index])
    error_index = list(set(index_mass) - set(n_error_index))
    return mass_translate, error_index


def write_text_file(mass_translate):
    with open('translate_file.txt', 'w') as f:
        # print(mass_translate)
        mass_translate = sorted(mass_translate)
        sring_translate = '; '.join(mass_translate)
        # print(sorted(sring_translate.split('; ')))
        print(sring_translate)
        f.write(sring_translate)


if __name__ == '__main__':
    name_file = 'exemple_class.txt'
    mass_text_ru = read_file(name_file)
    index_exel_ru, error_elements = search_index_element(mass_text_ru)
    mass_translate, error_index_en = search_translate(index_exel_ru)
    write_text_file(mass_translate)
    print('Количество русских элементов: ', len(mass_text_ru), '\n1)Количество индексов русских эл-ов: ', index_exel_ru,
          '\nОшибки поиска index - rus: ', error_elements, '\nОшибки поиска index - en',  error_index_en,
          '\nКоличество англиских элементов: ', len(mass_translate))
